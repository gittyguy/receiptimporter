"""
Excel writer.

Creates or updates the Expenses.xlsx workbook.

Columns

A : Purchase Date
B : Product
C : Total Price (Excel formula)
D : Shop
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook

from models import ReceiptDocument, ReceiptItem


class ExcelWriter:
    """
    Handles writing receipt items to Excel.
    """

    HEADERS = [
        "Purchase Date",
        "Product",
        "Total Price",
        "Shop",
    ]

    def __init__(self, filename: Path):

        self.filename = filename

        if self.filename.exists():

            self.workbook = load_workbook(self.filename)

            self.sheet = self.workbook.active

        else:

            self.workbook = Workbook()

            self.sheet = self.workbook.active

            self.sheet.title = "Expenses"

            self._create_headers()

            self._format_sheet()

    def _create_headers(self) -> None:

        self.sheet.append(self.HEADERS)

    def _format_sheet(self) -> None:

        self.sheet.column_dimensions["A"].width = 16
        self.sheet.column_dimensions["B"].width = 55
        self.sheet.column_dimensions["C"].width = 18
        self.sheet.column_dimensions["D"].width = 15

    def add_receipt(
        self,
        receipt: ReceiptDocument,
    ) -> None:
        """
        Add every product of a receipt.
        """

        for item in receipt.items:

            self.add_item(item)

    def add_item(
        self,
        item: ReceiptItem,
    ) -> None:
        """
        Append one product.
        """

        self.sheet.append([
            item.date,
            item.product,
            "",
            item.shop,
        ])

        row = self.sheet.max_row

        self.sheet[f"C{row}"] = self._build_formula(item)

    def _build_formula(
        self,
        item: ReceiptItem,
    ) -> str:
        """
        Build the Excel formula.

        If the receipt contains the printed line total,
        use that value because it avoids rounding errors.

        Otherwise calculate

            quantity × unit price

        Finally subtract the discount.
        """

        if item.line_total is not None:

            formula = f"={item.line_total:.2f}"

        else:

            formula = (
                f"={item.qty}"
                f"*{item.unit_price:.2f}"
            )

        if item.discount > 0:

            formula += f"-{item.discount:.2f}"

        return formula

    def save(self) -> None:

        self.workbook.save(self.filename)